import os
from collections import defaultdict
from datetime import datetime
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from sqlalchemy import select

from ..constants import (
    STATUS_ADDED,
    STATUS_ANSWER_RECEIVED,
    STATUS_CLARIFY,
    STATUS_DELETED_BY_USER,
    STATUS_IN_QUEUE,
    STATUS_LIST,
    STATUS_NEW,
    STATUS_NOT_ADDED,
)
from ..context import get_session_factory, get_settings
from ..models import Order, KindKeyword
from ..utils.photos import parse_photo_entries
from .photos import restore_order_photos


async def generate_order_reports(tmp_dir: str) -> Tuple[str, str]:
    settings = get_settings()
    session_factory = get_session_factory()
    async with session_factory() as session:
        q = await session.execute(select(Order).order_by(Order.created_at.asc()))
        rows = q.scalars().all()
        kw_rows = await session.execute(select(KindKeyword.kind, KindKeyword.keyword))
        keywords_map: Dict[str, List[str]] = defaultdict(list)
        for kind, kw in kw_rows.all():
            keywords_map[kind].append(kw)

    DEFAULT_KEYWORDS = {
        "Одежда": ["футболк", "куртк", "штаны", "джинс", "толстовк", "худи", "шорт", "юбк", "плать", "носки"],
        "Обувь": ["кроссовк", "ботин", "кеды", "сланс", "сланц", "сандал", "крос"],
        "Инвентарь": ["мяч", "гантел", "штанг", "скакалк", "коврик", "инвент"],
        "Аксессуары": ["рюкзак", "сумк", "перчат", "шапк", "шлем", "очки", "аксессуар"],
    }
    if not keywords_map:
        keywords_map.update(DEFAULT_KEYWORDS)
    KIND_VALUES = sorted(set(keywords_map.keys()) | set(DEFAULT_KEYWORDS.keys()))

    def guess_kind(name: str) -> str:
        title = (name or "").lower()
        for kind, keys in keywords_map.items():
            if any(k in title for k in keys):
                return kind
        return ""

    timestamp_human = datetime.utcnow().strftime("%d-%m-%Y %H-%M")
    full_path = os.path.join(tmp_dir, f"Все заказы {timestamp_human}.xlsx")
    work_path = os.path.join(tmp_dir, f"В работе {timestamp_human}.xlsx")

    def build_photo_columns(raw: str) -> Tuple[str, str]:
        entries = parse_photo_entries(raw, settings)
        locals_joined = "\n".join(local for local, _ in entries)
        public_joined = "\n".join(public for _, public in entries)
        return locals_joined, public_joined

    data_full: List[Dict[str, str]] = []
    for order in rows:
        await restore_order_photos(order.id)
        local_photos, public_photos = build_photo_columns(order.photos or "")
        data_full.append(
            {
                "ID заказа": order.id,
                "ID пользователя": order.user_id,
                "Статус": order.status,
                "Дата создания": order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
                "Товар": order.product,
                "Вид": guess_kind(order.product),
                "Бренд": order.brand,
                "Размер": order.size,
                "Комментарий": order.comment,
                "Фото (локально)": local_photos,
                "Ссылки на фото": public_photos,
                "Ссылка на товар": order.product_link,
                "Общение": order.communication,
                "Внутренние комментарии": order.internal_comments,
            }
        )
    df_full = pd.DataFrame(data_full)
    with pd.ExcelWriter(full_path, engine="openpyxl") as writer:
        df_full.to_excel(writer, index=False, sheet_name="Все заявки")
        ws_full = writer.sheets["Все заявки"]
        ws_full.freeze_panes = "A2"
        ws_status_full = writer.book.create_sheet("Статусы (полный)")
        for i, status in enumerate(STATUS_LIST, start=1):
            ws_status_full.cell(row=i, column=1, value=status)
        ws_status_full.sheet_state = "hidden"
        status_range_full = f"'Статусы (полный)'!$A$1:$A${len(STATUS_LIST)}"
        dv_full = DataValidation(type="list", formula1=status_range_full, allow_blank=False)
        dv_full.showErrorMessage = True
        dv_full.errorTitle = "Недопустимый статус"
        dv_full.error = "Выберите статус из списка."
        ws_full.add_data_validation(dv_full)
        dv_full.add(f"C2:C{len(df_full)+1}")

        # Валидация для столбца "Вид"
        kinds_sheet = writer.book.create_sheet("Справочник видов")
        for i, k in enumerate(KIND_VALUES, start=1):
            kinds_sheet.cell(row=i, column=1, value=k)
        kinds_sheet.sheet_state = "hidden"
        kind_range = f"'Справочник видов'!$A$1:$A${len(KIND_VALUES)}"
        dv_kind = DataValidation(type="list", formula1=kind_range, allow_blank=True)
        dv_kind.errorTitle = "Недопустимое значение"
        dv_kind.error = "Выберите значение из списка."
        ws_full.add_data_validation(dv_kind)
        dv_kind.add(f"F2:F{len(df_full)+1}")

        # Условное форматирование строк по статусу (приглушённые цвета)
        max_row = len(df_full) + 1
        status_rules = [
            (STATUS_ADDED, "BDE7BD"),
            (STATUS_NOT_ADDED, "F6D6D6"),
            (STATUS_CLARIFY, "E5E2F7"),
            (STATUS_NEW, "D9EAF7"),
            (STATUS_IN_QUEUE, "ECE9D8"),
            (STATUS_ANSWER_RECEIVED, "ECE9D8"),
            (STATUS_DELETED_BY_USER, "F2E2D2"),
        ]
        for status, color in status_rules:
            formula = f'$C2="{status}"'
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws_full.conditional_formatting.add(
                f"A2:N{max_row}",
                FormulaRule(formula=[formula], fill=fill),
            )

    data_work: List[Dict[str, str]] = []
    for order in rows:
        if order.status in (STATUS_ADDED, STATUS_NOT_ADDED):
            continue
        await restore_order_photos(order.id)
        local_photos, public_photos = build_photo_columns(order.photos or "")
        data_work.append(
            {
                "ID заказа": order.id,
                "ID пользователя": order.user_id,
                "Статус": order.status,
                "Дата создания": order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
                "Товар": order.product,
                "Вид": guess_kind(order.product),
                "Бренд": order.brand,
                "Размер": order.size,
                "Комментарий": order.comment,
                "Фото (локально)": local_photos,
                "Ссылки на фото": public_photos,
                "Ссылка на товар": order.product_link,
                "Общение": order.communication,
                "Внутренние комментарии": order.internal_comments,
            }
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Рабочий лист"
    headers = [
        "ID заказа",
        "ID пользователя",
        "Статус",
        "Дата создания",
        "Товар",
        "Вид",
        "Бренд",
        "Размер",
        "Комментарий",
        "Фото (локально)",
        "Ссылки на фото",
        "Ссылка на товар",
        "Общение",
        "Внутренние комментарии",
    ]
    ws.append(headers)
    for row in data_work:
        ws.append([row.get(h, "") for h in headers])

    ws_status = wb.create_sheet("Статусы")
    for i, status in enumerate(STATUS_LIST, start=1):
        ws_status.cell(row=i, column=1, value=status)
    ws_status.sheet_state = "hidden"
    status_range = f"'Статусы'!$A$1:$A${len(STATUS_LIST)}"
    dv = DataValidation(type="list", formula1=status_range, allow_blank=False)
    dv.showErrorMessage = True
    dv.errorTitle = "Недопустимый статус"
    dv.error = "Выберите статус из списка."
    ws.add_data_validation(dv)
    dv.add("C2:C1048576")

    # Валидация для столбца "Вид" (рабочий лист)
    kinds_sheet = wb.create_sheet("Справочник видов")
    for i, k in enumerate(KIND_VALUES, start=1):
        kinds_sheet.cell(row=i, column=1, value=k)
    kinds_sheet.sheet_state = "hidden"
    kind_range = f"'Справочник видов'!$A$1:$A${len(KIND_VALUES)}"
    dv_kind = DataValidation(type="list", formula1=kind_range, allow_blank=True)
    dv_kind.errorTitle = "Недопустимое значение"
    dv_kind.error = "Выберите значение из списка."
    ws.add_data_validation(dv_kind)
    dv_kind.add("F2:F1048576")

    # Условное форматирование строк по статусу
    status_rules = [
        (STATUS_ADDED, "BDE7BD"),
        (STATUS_NOT_ADDED, "F6D6D6"),
        (STATUS_CLARIFY, "E5E2F7"),
        (STATUS_NEW, "D9EAF7"),
        (STATUS_IN_QUEUE, "ECE9D8"),
        (STATUS_ANSWER_RECEIVED, "ECE9D8"),
        (STATUS_DELETED_BY_USER, "F2E2D2"),
    ]
    max_row = ws.max_row
    for status, color in status_rules:
        formula = f'$C2="{status}"'
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.conditional_formatting.add(
            f"A2:N{max_row}",
            FormulaRule(formula=[formula], fill=fill),
        )
    wb.save(work_path)

    return full_path, work_path


async def prepare_status_updates(path: str) -> Tuple[List[str], Dict[int, Dict[str, str]]]:
    errors: List[str] = []
    updates: Dict[int, Dict[str, str]] = {}

    try:
        df = pd.read_excel(path)
    except Exception as exc:
        errors.append(f"Не удалось прочитать файл: {exc}")
        return errors, updates

    required = {"ID заказа", "Статус"}
    missing = required - set(df.columns)
    if missing:
        errors.append(f"Отсутствуют столбцы: {', '.join(missing)}")
        return errors, updates

    if df["ID заказа"].duplicated().any():
        errors.append("Найдены дубликаты ID.")
        return errors, updates

    session_factory = get_session_factory()
    async with session_factory() as session:
        q = await session.execute(select(Order.id))
        existing = {row[0] for row in q.all()}

    for _, row in df.iterrows():
        try:
            order_id = int(row.get("ID заказа"))
        except Exception:
            errors.append(f"Некорректный ID: {row.get('ID заказа')}")
            continue
        if order_id not in existing:
            errors.append(f"Заказ {order_id} не найден.")
            continue
        new_status = str(row.get("Статус") or "").strip()
        if new_status not in STATUS_LIST:
            errors.append(f"Недопустимый статус для заказа {order_id}: {new_status}")
            continue
        product_link = str(row.get("Ссылка на товар") or "").strip()
        if new_status == STATUS_ADDED:
            link_lower = product_link.lower()
            if not (link_lower.startswith("http://") or link_lower.startswith("https://")):
                errors.append(f"Для заказа {order_id} нужен корректный URL товара.")
                continue
            if "https://www.sportmaster" not in link_lower:
                errors.append(f"Для заказа {order_id} ссылка должна вести на https://www.sportmaster.")
                continue
        updates[order_id] = {"status": new_status, "product_link": product_link}

    return errors, updates
